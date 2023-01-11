from vehicle_routing_problem.graph import Graph
import openpyxl


def parse_graph(file_path: str, cities: int) -> Graph:
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    graph = Graph()

    # cities + 2 because in excel indexing starts from 1 and A1 is empty since this is table
    for i in range(2, cities + 2):
        from_city = sheet.cell(row = 1, column = i).value
        for j in range(i + 1, cities + 2):
            to_city = sheet.cell(row = j, column = 1).value
            distance = float(sheet.cell(row = j, column = i).value)
            graph.add_edge(from_city, to_city, distance)
   
    return graph
